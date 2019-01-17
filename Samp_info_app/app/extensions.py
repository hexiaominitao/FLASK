import pandas as pd
from pandas import DataFrame
import zipfile, shutil, os, csv, re, openpyxl
from openpyxl import Workbook

from flask_bcrypt import Bcrypt
from flask_login import LoginManager
from flask_principal import Principal, Permission, RoleNeed
from flask_admin import Admin
from flask_cache import Cache
from flask_uploads import UploadSet, DOCUMENTS, DATA, ARCHIVES
from flask_celery import Celery
from flask_mail import Mail, Message

bcrypt = Bcrypt()
login_manager = LoginManager()
principal = Principal()
admin = Admin()
cache = Cache()
celery = Celery()
mail = Mail()

file_sample_info = UploadSet('filesam', DOCUMENTS)  # 文件上传
file_fastq_qc = UploadSet('filefastq', DATA)
file_bam_qc = UploadSet('filebam', DOCUMENTS)
file_zip = UploadSet('filezip', ARCHIVES)

login_manager.login_view = "main.login"
login_manager.session_protection = "strong"
login_manager.login_message = "Please login to access this page"
login_manager.login_message_category = "info"

admin_permission = Permission(RoleNeed('admin'))  # 添加权限 与manage 对应
poster_permission = Permission(RoleNeed('poster'))
default_permission = Permission(RoleNeed('default'))


@login_manager.user_loader
def load_user(user_id):
    from .models import User
    return User.query.get(user_id)


def excel_rd(path_xl):
    data_sample = pd.read_excel(path_xl, sep='/t', encoding='utf-8')
    df = DataFrame(data_sample)
    return df


def file_to_df(file):
    data_qc = []
    index_qc = []
    with open(file, 'r')as f1:
        f2 = csv.reader(f1, delimiter='\t')
        title = next(f2)
        for row in f2:
            sample_name_list = (row[title.index("#Sample_name")].split('/')[-1]).split('_')
            runname = (row[title.index("#Sample_name")].split('/')[-3])
            sample_name = '{}_{}'.format(sample_name_list[0], sample_name_list[2])
            index_qc.append(sample_name)
            data_qc.append(row[1:])
    df = DataFrame(data_qc, index=index_qc, columns=title[1:], dtype=int)
    df['Reads_p'] = df['Reads_num'] / df['Reads_num'].sum()
    df['Runname'] = runname

    return df, index_qc


def unzip_file(path_report, path_zip, path_vcf, filename, report_id):
    path_file = os.path.join(path_report, report_id)
    if not os.path.exists(path_file):
        os.mkdir(path_file)
    for name_zip in os.listdir(path_zip):
        if not name_zip.endswith('.zip'):
            continue
        if filename in name_zip:
            path_unzip = os.path.join(path_vcf, name_zip.strip('.zip'))
            zipfile.ZipFile(os.path.join(path_zip, name_zip)).extractall(path_unzip)
    for zip_root, zip_dir, zip_name in os.walk(path_unzip):
        for file_zip in zip_name:
            if file_zip.endswith('vcf') and 'Non-Filtered' in file_zip:
                # print(file_zip)
                shutil.copy2(os.path.join(zip_root, file_zip), path_file)
            elif file_zip.endswith('full.tsv'):
                tsvfile = os.path.join(zip_root, file_zip)
                shutil.copy2(os.path.join(zip_root, file_zip), path_file)
            elif file_zip.endswith('QC.pdf'):
                # print(file_zip)
                shutil.copy2(os.path.join(zip_root, file_zip), path_file)
    return tsvfile


def ir10087(report_id, file, path_report):
    wb = Workbook()
    dest_filename = report_id + '.xlsx'
    ws_ir = wb.active
    ws_ir.title = "IR"
    ws_mutation = wb.create_sheet(title="IR过滤")
    ws_cnv = wb.create_sheet(title="CNV")
    ws_fusion = wb.create_sheet(title="Fusion")
    ws_report = wb.create_sheet(title="REPORT")
    ws_cov = wb.create_sheet(title="覆盖度")
    ws_details = wb.create_sheet(title="详情")
    csvRows = []
    out_ir = []
    out_Mutation = []
    out_CNV = []
    out_Fusion = []
    path_wk = os.getcwd()
    path_file = os.path.join(path_wk, path_report, report_id)
    file_vcf = os.path.join(path_wk, file)
    os.chdir(path_file)

    def write_tsv(WS, TSV):
        for row_1 in range(len(TSV)):
            WS.append(TSV[row_1])

    with open(file_vcf, 'r')as f_csv:
        f_read = csv.reader(f_csv, delimiter='\t')
        next(f_read)
        next(f_read)
        for row in f_read:
            csvRows.append(row)
    title_ir = csvRows[0]
    af, fao = '', ''
    for i in range(1, len(csvRows)):
        allele_i = csvRows[i][csvRows[0].index('allele_coverage')].split(',')
        allele = allele_i[0]
        coverage = csvRows[i][csvRows[0].index('coverage')]
        fao_2, af_2 = [], []
        if allele.strip():
            if int(coverage) == 0:
                af, fao = ' ', ' '
            else:
                for j in range(1, len(allele_i)):
                    if float(allele_i[j]):
                        fao_1 = allele_i[j]
                        af_1 = str(round((float(allele_i[j]) / float(coverage) * 100),2)) + '%'
                        fao_2.append(fao_1)
                        af_2.append(af_1)
                        fao = ','.join(fao_2)
                        af = ','.join(af_2)
                    else:
                        af, fao = str(round((100 - float(allele) / float(coverage) * 100), 2)) + '%', int(
                            coverage) - int(
                            allele)
        else:
            af, fao = ' ', ' '
        csvRows[i].insert(title_ir.index('coverage'), fao)
        csvRows[i].insert(title_ir.index('coverage'), af)
        out_ir.append(csvRows[i])
    title_ir.insert(title_ir.index('coverage'), 'AF')
    title_ir.insert(title_ir.index('coverage'), 'FAO')
    out_ir.insert(0, title_ir)
    # file_open_w(ir_name + 'IR.csv', out_ir)
    write_tsv(ws_ir, out_ir)
    title = out_ir[0]
    for j in range(len(out_ir)):
        ir = out_ir[j]
        Fu = ir[title.index('function')]
        m = re.search(
            '^(missense|nonsense|nonframeshiftDeletion|nonframeshiftInsertion|frameshiftDeletion'
            '|frameshiftInsertion '
            '|frameshiftBlockSubstitution|nonframeshiftBlockSubstitution)',
            Fu)
        fusion = ['ASSAYS_5P_3P', 'EXPR_CONTROL', 'FUSION']
        if m is not None:

            if 0.00 < float(ir[title.index('AF')].split(',')[0].strip('%')):
                out_Mutation.append(ir)
        elif ir[title.index('type')] in fusion:
            if ir[title.index('fusion_presence')] == "Present":
                out_Fusion.append(ir)
        elif ir[title.index('type')] == "CNV":
            Cnv = ir[title.index('confidence')].split(',')
            Cnv_ad = Cnv[0].split(':')
            Cnv_re = Cnv[1].split(':')
            if float(Cnv_ad[1]) >= 4:
                out_CNV.append(ir)
            if float(Cnv_re[1]) <= 1:
                out_CNV.append(ir)
    if out_Mutation:
        out_Mutation.insert(0, title)
        # file_open_w(ir_name + 'Mutation.csv', out_Mutation)
        write_tsv(ws_mutation, out_Mutation)
        # print('检测到了%d个DNA序列变异,详情见 IR过滤 ' % (len(out_Mutation) - 1))
    else:
        # print('未检测到突变,请用IGV手工查看')
        pass
    if out_Fusion:
        out_Fusion.insert(0, title)
        # file_open_w(ir_name + 'Fusion.csv', out_Fusion)
        write_tsv(ws_fusion, out_Fusion)
        # print('融合结果请查看 Fusion ')
    else:
        pass
        # print('这个文件没有融合信息')

    if out_CNV:
        out_CNV.insert(0, title)
        # file_open_w(ir_name + 'CNV.csv', out_CNV)
        write_tsv(ws_cnv, out_CNV)
        # print("检测到了%d个DNA拷贝数变异，详情见 CNV" % (len(out_CNV) - 1))
    else:
        pass
        # print("未检测到DNA拷贝数变异")

    out_Report = []
    out_Result = ''
    title_Report = ['基因', '突变类型', '突变名称', '突变全称', '突变频率', '覆盖度']
    for k in range(1, len(out_Mutation)):
        ir_m = out_Mutation[k]
        Protein, Function = ir_m[title.index("protein")].split('|')[0], \
                            ir_m[title.index("function")].split('|')[0]
        Re_GENE, De_GENE = [], []
        if ir_m[title.index("function")].split('|')[0] in ["nonframeshiftDeletion",
                                                           "frameshiftDeletion"]:
            GENE = ir_m[title.index("coding")].split('|')[0].split('del')[-1]
            Re_GENE, De_GENE = GENE, GENE + '/-'
        elif ir_m[title.index("function")].split('|')[0] in ['nonsense', 'missense']:
            GENE = ir_m[title.index("coding")].split('|')[0][-3:].split('>')
            Re_GENE, De_GENE = GENE[0], GENE[0] + "/" + GENE[-1]
        elif ir_m[title.index("function")].split('|')[0] in ['frameshiftInsertion',
                                                             'nonframeshiftInsertion']:
            GENE = ir_m[title.index("coding")].split('|')[0].split('ins')[-1]
            Re_GENE, De_GENE = '-', '-/' + GENE

        rep_P = {"Ala": "A", "Arg": "R", "Asn": "N", "Asp": "D", "Cys": "C", "Gln": "Q", "Glu": "E",
                 "Gly": "G",
                 "His": "H",
                 "Ile": "I", "Leu": "L", "Lys": "K", "Met": "M", "Phe": "F", "Pro": "P", "Ser": "S",
                 "Thr": "T",
                 "Trp": "W",
                 "Tyr": "Y", "Val": "V", "p.": "", }
        rep_F = {'missense': '错义突变', 'nonsense': '无义突变', 'nonframeshiftDeletion': '非移码缺失突变',
                 'frameshiftDeletion': '移码缺失突变', 'nonframeshiftInsertion': '非移码插入突变',
                 'frameshiftInsertion': '移码插入突变'}

        rep_P, rep_F = dict((re.escape(k1), v1) for k1, v1 in rep_P.items()), dict(
            (re.escape(k2), v2) for k2, v2 in rep_F.items())

        pattern_P, pattern_F = re.compile("|".join(rep_P.keys())), re.compile("|".join(rep_F.keys()))

        Protein, Function = pattern_P.sub(lambda n: rep_P[re.escape(n.group(0))],
                                          Protein), pattern_F.sub(
            lambda n: rep_F[re.escape(n.group(0))], Function)

        Name = ir_m[title.index("gene")].split('|')[0] + ' ' + Protein
        Name_a = ir_m[title.index('transcript')].split('|')[0] + '(' + ir_m[title.index("gene")].split('|')[
            0] + ')' \
                 + ':' + ir_m[title.index("coding")].split('|')[0] + ' (' + \
                 ir_m[title.index('protein')].split('|')[0] + ')'

        Result = ir_m[title.index("gene")].split('|')[0] + ' ' + Protein + '(' + str(
            round(float(ir_m[title.index('AF')].split(',')[0].strip('%')), 2)) + '%),'

        report_AF = ir_m[title.index('AF')].split(',')[0]
        report = [ir_m[title.index("gene")].split('|')[0], Function, Name, Name_a,
                  report_AF, ir_m[title.index('FAO')]]
        out_Report.append(report)
        out_Result = out_Result + Result
    # print('报告结果>>>>>> %s' % out_Result)
    if out_Report:
        out_Report.insert(0, title_Report)
        # file_open_w(ir_name + 'REPORT.csv', out_Report)
        write_tsv(ws_report, out_Report)
        # print('结果详见 REPORT（有缺失突变 或插入突变请查看IGV获得 <参考基因型>和<检测基因型> ）')
        ws_details['A5'] = out_Result
    wb.save(filename=dest_filename)
    os.chdir(path_wk)
    return out_Report


def archive_file(path_report, filename):
    path_wk = os.getcwd()
    with zipfile.ZipFile('{}/{}.zip'.format(path_report, filename), 'w') as myzip:
        os.chdir(path_report)
        for arc_root, arc_dir, arc_file in os.walk(filename):
            for file in arc_file:
                myzip.write(os.path.join(arc_root, file), file)
    os.chdir(path_wk)
